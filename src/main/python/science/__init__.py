import os
from io import BytesIO

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5 import FigureCanvasQT

from openpyxl import load_workbook

FACTORS = [
    "Без нагрузки",
    "С физической нагрузкой",
    "С эмоциональной нагрузкой",
    "После отдыха"
]
FACTORS_L = [factor.lower() for factor in FACTORS]


class XLSXParseError(ValueError):
    pass


def is_float(s: str):
    try:
        return float(s)
    except ValueError:
        return False


def file_base_name(filename: str):
    return os.path.splitext(os.path.basename(filename))[0]


def read_sample(filename):
    """Чтение эталонов"""
    with open(filename) as file:
        data = [row.strip() for row in file]

    for idx, el in enumerate(data):
        # noinspection PyTypeChecker
        data[idx] = float(el)
    return data


def read_xlsx_sample(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.get_active_sheet()

    datas = []
    for col in range(1, 4 + 1):
        data, row = [], 1
        val = sheet.cell(row, col).value
        if not val:
            datas.append(None)
            continue

        val = val.strip()
        # Пропустим первую строку, если там не число (заголовок, например)
        if is_float(val) is False:
            row += 1
        while sheet.cell(row, col).value is not None and sheet.cell(row, col).value.strip():
            try:
                data.append(float(sheet.cell(row, col).value))
            except ValueError:
                err = 'Ошибка при парсе файла {}, страницы {}, ячейки {}{}'.format(filename,
                                                                                   sheet.title,
                                                                                   'ABCD'[col],
                                                                                   row)
                print(err)
                raise XLSXParseError(err)
            row += 1
        datas.append(data)
    return datas


def patient_suffix(filename: str, suffix):
    return filename[:filename.rfind('.')] + suffix + filename[filename.rfind('.'):]


def plot_image(plot_func, *args, **kwargs):
    figure = Figure(dpi=100)
    canvas = FigureCanvasQT(figure)
    plot_func(*args, figure)

    buffer = BytesIO()
    canvas.print_figure(buffer)
    if kwargs.get('io', False):
        buffer.seek(0)
        return buffer
    else:
        return buffer.getvalue()


if __name__ == '__main__':
    print('\n'.join(map(str, read_xlsx_sample("samples/1_1.xlsx"))))

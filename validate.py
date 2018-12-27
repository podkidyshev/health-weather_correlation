import os
import sys
from openpyxl import load_workbook


def is_float(s: str):
    try:
        return float(s)
    except ValueError:
        return False


def read_xlsx_sample(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.get_active_sheet()

    datas = []
    for col in range(1, 4 + 1):
        data, row = [], 1
        val = sheet.cell(row, col).value
        if not val:
            datas.append(None)
            continue

        val = val.strip()
        # Пропустим первую строку, если там не число (заголовок, например)
        if is_float(val) is False:
            row += 1
        while sheet.cell(row, col).value is not None and sheet.cell(row, col).value.strip():
            try:
                data.append(float(sheet.cell(row, col).value))
            except ValueError:
                err = 'Ошибка при парсе файла {}, страницы {}, ячейки {}{}'.format(filename,
                                                                                   sheet.title,
                                                                                   'ABCD'[col],
                                                                                   row)
                print(err)
            row += 1
        datas.append(data)
    return datas


def read_sample(filename):
    """Чтение эталонов"""
    with open(filename) as file:
        data = [row.strip() for row in file]

    for idx, el in enumerate(data):
        # noinspection PyTypeChecker
        data[idx] = float(el)
    return data


# with open(r"source_original\Анализ 06.12.18\2_2o.txt") as f:
#     print(f.read())
# print(read_sample(r"source_original\Анализ 06.12.18\2_2o.txt"))
# exit(0)


path_test = sys.argv[1]
path_orig = sys.argv[2]

for group in '123':
    for idx in '123456':
        name = group + '_' + idx
        datas = read_xlsx_sample(os.path.join(path_test, name + '.xlsx'))

        datas_orig = []
        for suffix in ['', 'n', 'e', 'o']:
            datas_orig.append(read_sample(os.path.join(path_orig, name + suffix + '.txt')))

        for idx, test, orig in zip(range(4), datas, datas_orig):
            if test is None:
                print('НОНЕ,', name + '_' + str(idx))
            elif len(test) != len(orig):
                print('ДЛИНЫ,', name + '_' + str(idx))
                # print(os.path.join(path_orig, name + suffix + '.txt'))
                # print(os.path.join(path_test, name + '.xlsx'))
                # print(test)
                print(orig)
            else:
                for el_test, el_orig in zip(test, orig):
                    if el_test != el_orig:
                        print('ЗНАЧЕНИЯ,', name + suffix)
                        break

